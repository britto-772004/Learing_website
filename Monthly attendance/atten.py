import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def update_attendance_for_all_sessions(file_path, roll_numbers, date_column):
    # Load the Excel workbook and select the first sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Find all column indices for the given date
    date_column_indices = []
    for col in range(4, sheet.max_column + 1):  # Start searching from the 4th column
        if str(sheet.cell(row=1, column=col).value).startswith(date_column):
            date_column_indices.append(col)

    if not date_column_indices:
        st.error("Date not found in the sheet!")
        return None

    # Update attendance for each roll number
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for roll_number in roll_numbers:
        # Find the row for the given roll number
        roll_number_row_index = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == roll_number:
                roll_number_row_index = row
                break

        if not roll_number_row_index:
            st.warning(f"Roll number {roll_number} not found in the sheet!")
            continue

        # Update the attendance to "A" (Absent) for all sessions on the given date
        for col in date_column_indices:
            sheet.cell(row=roll_number_row_index, column=col).value = "A"
            sheet.cell(row=roll_number_row_index, column=col).fill = red_fill

    # Save the updated workbook to a new file
    output_path = "Updated_Student_Attendence_AIB.xlsx"
    workbook.save(output_path)
    return output_path

# Streamlit UI
st.title("Attendance Updater")

# Directly reference the Excel file from a predefined path
file_path = "C:/Users/VIKRAM RAAJA K/MY WORK/Class/Monthly attendance/AIB.xlsx"  # Replace this with the actual path to your Excel file

# Default prefix
roll_number_prefix = "22AIB"

# Option 1: Input roll number suffixes as a comma-separated list
roll_number_suffixes_input = st.text_area("Enter Roll Number Suffixes (comma-separated, e.g., 01, 02, 03)", "01, 02")

# Convert the suffix input to a list and generate full roll numbers
roll_numbers = [f"{roll_number_prefix}{suffix.strip()}" for suffix in roll_number_suffixes_input.split(",") if suffix.strip()]

# Option 2: Upload a text file with roll number suffixes (one per line)
roll_numbers_file = st.file_uploader("Or upload a text file with Roll Number Suffixes", type="txt")

if roll_numbers_file:
    # Read suffixes from file and generate full roll numbers
    roll_numbers = [f"{roll_number_prefix}{line.strip()}" for line in roll_numbers_file.readlines()]

date_column = st.text_input("Enter Date (e.g., 2024-08-07)", "2024-08-07")

# Update attendance on button click
if st.button("Update Attendance"):
    if roll_numbers:
        output_file = update_attendance_for_all_sessions(file_path, roll_numbers, date_column)
        if output_file:
            with open(output_file, "rb") as file:
                btn = st.download_button(
                    label="Download Updated Excel",
                    data=file,
                    file_name="UpdatedAIB.xlsx"
                )
    else:
        st.error("Please provide at least one roll number suffix.")
