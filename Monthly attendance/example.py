#It is for a full day.
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

# Function to get available dates from the Excel sheet
def get_available_dates(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    available_dates = []

    # Iterate through the first row to extract dates
    for col in range(4, sheet.max_column + 1, 8):  # Assuming dates start from the 4th column and are spaced 8 columns apart
        cell_value = sheet.cell(row=1, column=col).value
        try:
            current_date = datetime.strptime(str(cell_value).split()[0], "%Y-%m-%d").date()
            available_dates.append(current_date)
        except (ValueError, TypeError):
            continue

    return available_dates

# Function to update attendance for the whole day
def update_full_day_attendance(file_path, date_roll_numbers):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Red fill for "A" (Absent)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Iterate over the provided date and roll number data
    for date, roll_numbers in date_roll_numbers.items():
        # Find the column for the given date (first column of that date)
        date_col_index = None
        for col in range(4, sheet.max_column + 1, 8):  # Skip 8 columns for each date
            cell_value = sheet.cell(row=1, column=col).value
            try:
                current_date = datetime.strptime(str(cell_value).split()[0], "%Y-%m-%d").date()
                if current_date == date:
                    date_col_index = col  # This is the start column for this date
                    break
            except (ValueError, TypeError):
                continue

        if date_col_index is None:
            st.error(f"No matching date found in the sheet: {date}.")
            continue

        # Update attendance for the specified roll numbers for the entire day
        for roll_number in roll_numbers:
            # Find the row for the given roll number
            roll_number_row_index = None
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=2).value == roll_number:
                    roll_number_row_index = row
                    break

            if not roll_number_row_index:
                st.warning(f"Roll number {roll_number} not found in the sheet!")
                continue

            # Update the attendance to "A" for all hours of the specific date
            for hour in range(1, 9):  # Assuming 8 hours per day
                sheet.cell(row=roll_number_row_index, column=date_col_index + (hour - 1)).value = "A"
                sheet.cell(row=roll_number_row_index, column=date_col_index + (hour - 1)).fill = red_fill  # Apply red fill

    # Save the updated workbook to a new file
    output_path = "Updated_Attendance_Full_Day_AIB.xlsx"
    workbook.save(output_path)
    return output_path

# Streamlit UI
st.title("Update Attendance for a Full Day")

# Directly reference the Excel file from a predefined path
file_path = "C:/Users/VIKRAM RAAJA K/MY WORK/Class/Monthly attendance/Student Attendence AIB.xlsx"  # Replace this with the actual path to your Excel file

# Display available dates from the Excel sheet
available_dates = get_available_dates(file_path)
st.subheader("Available Dates in the Excel Sheet:")
if available_dates:
    st.markdown("<ol style='list-style-type: decimal;'>", unsafe_allow_html=True)  # Start of ordered list
    for date in available_dates:
        st.markdown(f"<li>{date.strftime('%Y-%m-%d')}</li>", unsafe_allow_html=True)  # List items in the desired format
    st.markdown("</ol>", unsafe_allow_html=True)  # End of ordered list
else:
    st.write("No available dates found in the Excel sheet.")

# Default prefix for roll numbers
roll_number_prefix = "22AIB"

# Input number of days to update attendance for a full day
max_days = len(available_dates)  # Maximum number of days is equal to the available dates
num_days = st.number_input("How many days do you want to update attendance for?", min_value=1, max_value=max_days, value=1)

# Dictionary to hold date and corresponding roll numbers
date_roll_numbers = {}

for i in range(num_days):
    # Input date with auto-suggest
    if i == 0:
        date_input = st.selectbox(
            f"Select Date {i + 1}",
            options=[date.strftime('%Y-%m-%d') for date in available_dates],
            index=0 if available_dates else None,
            format_func=lambda x: x if x else "No available dates"
        )
        selected_date = datetime.strptime(date_input, "%Y-%m-%d").date()
    else:
        # Automatically select the next available date
        next_available_date = available_dates[available_dates.index(selected_date) + 1] if selected_date in available_dates and available_dates.index(selected_date) + 1 < len(available_dates) else None
        date_input = st.selectbox(
            f"Select Date {i + 1}",
            options=[date.strftime('%Y-%m-%d') for date in available_dates],
            index=available_dates.index(next_available_date) if next_available_date else 0,
            format_func=lambda x: x if x else "No available dates"
        )
        selected_date = datetime.strptime(date_input, "%Y-%m-%d").date()

    try:
        # Input the corresponding roll numbers for the date
        roll_numbers_input = st.text_area(f"Enter Roll Numbers for {selected_date} (comma-separated, e.g., 01, 02, 03)", "", key=f"roll_numbers_{i}")
        roll_numbers = [f"{roll_number_prefix}{suffix.strip()}" for suffix in roll_numbers_input.split(",") if suffix.strip()]

        if roll_numbers:
            date_roll_numbers[selected_date] = roll_numbers

    except ValueError:
        st.error(f"Invalid date format for Date {i + 1}. Please use 'yyyy-mm-dd' format.")

# Update attendance on button click
if st.button("Update Attendance for Full Day"):
    if date_roll_numbers:
        output_file = update_full_day_attendance(file_path, date_roll_numbers)
        if output_file:
            st.success("Attendance for the full day updated successfully!")
            with open(output_file, "rb") as file:
                st.download_button(
                    label="Download Updated Excel",
                    data=file,
                    file_name="UpdatedAttendanceFullDayAIB.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.error("Please provide valid dates and roll numbers.")
