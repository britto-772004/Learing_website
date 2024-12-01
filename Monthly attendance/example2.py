import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# Function to update specific hours of absence
def update_specific_hours(file_path, date_hours_roll_numbers):
    # Load the Excel workbook and select the first sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Red fill for "A" (Absent)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Iterate over the provided date, hours, and roll number data
    for date, hours_roll_numbers in date_hours_roll_numbers.items():
        # Find the column for the given date (first column of that date)
        date_col_index = None
        for col in range(4, sheet.max_column + 1, 8):  # Skip 8 columns for each date
            cell_value = sheet.cell(row=1, column=col).value
            try:
                current_date = datetime.strptime(str(cell_value).split()[0], "%Y-%m-%d").date()
                if current_date == date:
                    date_col_index = col  # This is the start column for this date
                    break
            except (ValueError, TypeError):
                continue

        if date_col_index is None:
            st.error(f"No matching date found in the sheet: {date}.")
            continue

        # Update attendance for the specified hours and roll numbers
        for hour, roll_numbers in hours_roll_numbers.items():
            for roll_number in roll_numbers:
                # Find the row for the given roll number
                roll_number_row_index = None
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=2).value == roll_number:
                        roll_number_row_index = row
                        break

                if not roll_number_row_index:
                    st.warning(f"Roll number {roll_number} not found in the sheet!")
                    continue

                # Update the attendance to "A" for the specific hour
                sheet.cell(row=roll_number_row_index, column=date_col_index + (hour - 1)).value = "A"
                sheet.cell(row=roll_number_row_index, column=date_col_index + (hour - 1)).fill = red_fill  # Apply red fill

    # Save the updated workbook to a new file
    output_path = "Updated_Attendance_Hours_AIB.xlsx"
    workbook.save(output_path)
    return output_path

# Streamlit UI
st.title("Update Attendance for Specific Hours")

# Directly reference the Excel file from a predefined path
file_path = "C:/Users/VIKRAM RAAJA K/MY WORK/Class/Monthly attendance/Student Attendence AIB.xlsx"  # Replace this with the actual path to your Excel file

# Default prefix for roll numbers
roll_number_prefix = "22AIB"

# Display available dates from the Excel sheet
def get_available_dates(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    available_dates = []

    # Iterate through the first row to extract dates
    for col in range(4, sheet.max_column + 1, 8):  # Assuming dates start from the 4th column and are spaced 8 columns apart
        cell_value = sheet.cell(row=1, column=col).value
        try:
            current_date = datetime.strptime(str(cell_value).split()[0], "%Y-%m-%d").date()
            available_dates.append(current_date)
        except (ValueError, TypeError):
            continue

    return available_dates

available_dates = get_available_dates(file_path)
st.subheader("Available Dates in the Excel Sheet:")
if available_dates:
    st.markdown("<ol style='list-style-type: decimal;'>", unsafe_allow_html=True)  # Start of ordered list
    for date in available_dates:
        st.markdown(f"<li>{date.strftime('%Y-%m-%d')}</li>", unsafe_allow_html=True)  # List items in the desired format
    st.markdown("</ol>", unsafe_allow_html=True)  # End of ordered list
else:
    st.write("No available dates found in the Excel sheet.")

# Input number of days to update attendance for specific hours
num_days = st.number_input("How many days do you want to update specific hours for?", min_value=1, max_value=31, value=1)

# Dictionary to hold date, hours, and corresponding roll numbers
date_hours_roll_numbers = {}

for i in range(num_days):
    # Input date with auto-suggest
    date_input = st.selectbox(
        f"Select Date {i + 1}",
        options=[date.strftime('%Y-%m-%d') for date in available_dates],
        index=0 if available_dates else None,
        format_func=lambda x: x if x else "No available dates"
    )
    try:
        date_obj = datetime.strptime(date_input, "%Y-%m-%d").date()

        # Input the specific hours and corresponding roll numbers for the date
        hours_roll_numbers = {}
        for hour in range(1, 9):  # Assuming 8 hours per day
            roll_numbers_input = st.text_area(
                f"Enter Roll Numbers for Hour {hour} on {date_obj} (comma-separated, e.g., 01, 02, 03)", 
                key=f"roll_numbers_hour_{hour}_date_{i}"  # Assign a unique key for each input
            )
            roll_numbers = [f"{roll_number_prefix}{suffix.strip()}" for suffix in roll_numbers_input.split(",") if suffix.strip()]

            if roll_numbers:
                hours_roll_numbers[hour] = roll_numbers

        # Store the hours and roll numbers for the date
        date_hours_roll_numbers[date_obj] = hours_roll_numbers

    except ValueError:
        st.error(f"Invalid date format for Date {i + 1}. Please use 'yyyy-mm-dd' format.")

# Update attendance on button click
if st.button("Update Attendance for Specific Hours"):
    if date_hours_roll_numbers:
        output_file = update_specific_hours(file_path, date_hours_roll_numbers)
        if output_file:
            st.success("Attendance for specific hours updated successfully!")
            with open(output_file, "rb") as file:
                st.download_button(
                    label="Download Updated Excel",
                    data=file,
                    file_name="UpdatedAttendanceHoursAIB.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.error("Please provide valid dates, hours, and roll numbers.")
