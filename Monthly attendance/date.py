import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import tempfile
from openpyxl.utils import get_column_letter

# Function to adjust the column width based on the length of the content in the cells
def adjust_column_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None  # Initialize a variable to store the column letter
        for cell in column_cells:
            # Check if the cell is part of a merged range
            if not any([cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges]):
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    if column_letter is None:
                        column_letter = get_column_letter(cell.column)  # Get the column letter (e.g., 'A', 'B', etc.)
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)  # Add some extra space for readability
            ws.column_dimensions[column_letter].width = adjusted_width


# Streamlit title
st.title("Student Details Input and Holiday Selector")

# Step 1: Upload Student Details Excel file
uploaded_file = st.file_uploader("Upload Student Details Excel", type=["xlsx"])

if uploaded_file is not None:
    # Read the student details from the uploaded Excel file
    df = pd.read_excel(uploaded_file)
    st.write("Uploaded Student Details:")
    st.dataframe(df)

    # Step 2: Holiday Selector and Working Days Generator
    st.title("Holiday Selector and Working Days Generator")

    # Select the start and end date for the range (working period)
    start_date = st.date_input("Select Start Date for the Working Period", datetime.now())
    end_date = st.date_input("Select End Date for the Working Period", datetime.now() + timedelta(days=30))

    if start_date > end_date:
        st.error("End date must be after start date!")
    else:
        # Generate the full list of dates in the given range
        full_working_period = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Allow users to select holidays from the generated range
        selected_holidays = st.multiselect("Select Holidays", full_working_period)

        # Calculate working days by removing the holidays
        working_days = [day for day in full_working_period if day not in selected_holidays]

        # Display the working days
        st.write("Working Days (excluding holidays):")
        for day in working_days:
            st.write(day)

        # Display the selected holidays
        if selected_holidays:
            st.write("Selected Holidays:")
            for holiday in selected_holidays:
                st.write(holiday)
        else:
            st.write("No holidays selected.")

        # Button to download the updated Excel sheet with student details and working days
        if st.button("Download Updated Excel"):
            # Load the existing student details Excel file
            wb = load_workbook(uploaded_file)
            ws = wb.active

            # Find the last row with data (after the student details)
            last_row = ws.max_row + 2  # Leave 1 row of space after the student details

            # Add the working days and holidays information after the student details
            ws.cell(row=last_row, column=1, value="Total Working Days:")
            ws.cell(row=last_row, column=2, value=len(working_days))

            ws.cell(row=last_row + 1, column=1, value="Total Holidays:")
            ws.cell(row=last_row + 1, column=2, value=len(selected_holidays))

            # Add the header row (dates) to the worksheet with merged columns
            for i, date in enumerate(working_days):
                col_start = (i * 8) + 4  # Start from the fourth column (A=1, B=2, C=3, D=4)
                col_end = col_start + 7  # Merge 8 columns for each date
                ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)

                # Set the merged cell value
                merged_cell = ws.cell(row=1, column=col_start, value=date.strftime("%Y-%m-%d"))

                # Center align the merged cell both horizontally and vertically
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Print numbers 1 to 8 below each date
                for j in range(1, 9):  # 1 to 8
                    ws.cell(row=2, column=col_start + (j - 1), value=j).alignment = Alignment(horizontal='center', vertical='center')

            # Adjust the width of all columns based on the content
            adjust_column_width(ws)

            # Save the updated Excel workbook to a temporary file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)

            # Offer the updated file for download
            with open(tmp.name, "rb") as file:
                st.download_button(
                    label="Download Updated Excel with Student Details",
                    data=file,
                    file_name="updated_student_details_and_working_days.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
